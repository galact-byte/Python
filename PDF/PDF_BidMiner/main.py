#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BidMiner - 招标文件信息提取工具
依赖: pymupdf dashscope openpyxl
"""

import os
import json
import argparse
import datetime
import base64
import time
import re
from pathlib import Path

import fitz  # pymupdf
import openpyxl
from http import HTTPStatus
import dashscope
from dashscope import Generation, MultiModalConversation

# ──────────────────────────────────────────
# 配置
# ──────────────────────────────────────────
TEXT_MODEL  = "qwen-turbo"
VISION_MODEL = "qwen-vl-max"
IMAGE_DPI   = 300          # 图片型PDF渲染分辨率，越高越准但越慢
MAX_RETRIES = 3             # API调用失败重试次数
RETRY_DELAY = 5             # 重试等待秒数

EXCEL_HEADERS = [
    "文件名", "信息类型", "项目名称", "开标时间", "城市",
    "行业", "服务内容", "预算（万元）",
    "采购人", "代理机构",
    "中标单位", "中标金额（万元）", "提取备注"
]


# ──────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────

def load_api_key(api_key_file: str) -> str:
    with open(api_key_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    key = data.get("DASHSCOPE_API_KEY", "")
    if not key or key.startswith("sk-xxx"):
        raise ValueError("请在 api_key.json 中填写有效的 DASHSCOPE_API_KEY")
    return key


def load_prompt(prompt_file: str) -> str:
    with open(prompt_file, "r", encoding="utf-8") as f:
        return f.read().strip()


def is_image_based_pdf(pdf_path: str, text_threshold: int = 50) -> bool:
    """判断PDF是否为图片型（正文内容在图片里）"""
    doc = fitz.open(pdf_path)
    total_text = ""
    for page in doc:
        total_text += page.get_text()
    doc.close()
    text = total_text.strip()
    # 明确标注"正文详见图片"的，强制走视觉模型
    if "正文详见图片" in text:
        return True
    # 文字少于阈值字符，认为是图片型
    return len(text) < text_threshold


def extract_text_from_pdf(pdf_path: str) -> str:
    """从文字型PDF提取文本"""
    doc = fitz.open(pdf_path)
    text_parts = []
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return "\n".join(text_parts).strip()


def pdf_pages_to_base64(pdf_path: str, dpi: int = IMAGE_DPI) -> list:
    """将PDF每页渲染为base64图片列表"""
    doc = fitz.open(pdf_path)
    images_b64 = []
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    for page in doc:
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        images_b64.append(base64.b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return images_b64


def parse_json_from_response(text: str) -> dict:
    """从模型返回文本中提取JSON"""
    text = text.strip()
    # 去掉可能的 markdown 代码块
    if "```" in text:
        lines = text.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        text = "\n".join(lines)
    # 找第一个 { 到最后一个 }
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1:
        text = text[start:end+1]
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {}


# ──────────────────────────────────────────
# API 调用
# ──────────────────────────────────────────

def call_text_model(prompt: str, content: str) -> dict:
    """调用文本模型提取信息"""
    messages = [
        {"role": "user", "content": prompt + "\n\n" + content}
    ]
    for attempt in range(MAX_RETRIES):
        try:
            response = Generation.call(
                model=TEXT_MODEL,
                messages=messages,
                result_format="message"
            )
            if response.status_code == HTTPStatus.OK:
                reply = response.output.choices[0].message.content
                return parse_json_from_response(reply)
            else:
                print(f"  [文本模型] 错误 {response.status_code}: {response.message}")
        except Exception as e:
            print(f"  [文本模型] 异常: {e}")
        if attempt < MAX_RETRIES - 1:
            print(f"  等待 {RETRY_DELAY}s 后重试...")
            time.sleep(RETRY_DELAY)
    return {}


def call_vision_model(prompt: str, images_b64: list) -> dict:
    """调用视觉模型处理图片型PDF，所有页合并为一次请求"""
    images_b64 = images_b64[:5]

    content = []
    if len(images_b64) > 1:
        content.append({"text": f"以下是一份PDF文件，共{len(images_b64)}页，请综合所有页面内容提取信息。"})
    for i, img_b64 in enumerate(images_b64):
        if len(images_b64) > 1:
            content.append({"text": f"第{i+1}页："})
        content.append({"image": f"data:image/png;base64,{img_b64}"})
    content.append({"text": prompt})

    messages = [{"role": "user", "content": content}]

    for attempt in range(MAX_RETRIES):
        try:
            response = MultiModalConversation.call(
                model=VISION_MODEL,
                messages=messages
            )
            if response.status_code == HTTPStatus.OK:
                reply = response.output.choices[0].message.content
                if isinstance(reply, list):
                    reply = " ".join([r.get("text", "") for r in reply])
                return parse_json_from_response(reply)
            else:
                print(f"  [视觉模型] 错误 {response.status_code}: {response.message}")
        except Exception as e:
            print(f"  [视觉模型] 异常: {e}")
        if attempt < MAX_RETRIES - 1:
            print(f"  等待 {RETRY_DELAY}s 后重试...")
            time.sleep(RETRY_DELAY)
    return {}


# ──────────────────────────────────────────
# Excel 输出
# ──────────────────────────────────────────

def init_excel(output_path: str) -> tuple:
    """初始化或加载Excel文件，返回 (workbook, worksheet)"""
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "招标信息"
        ws.append(EXCEL_HEADERS)
        # 表头样式：蓝色背景+白色加粗字体
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        # 设置列宽
        col_widths = [25, 16, 40, 18, 10, 10, 30, 14, 20, 20, 20, 14, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    return wb, ws


def normalize_text(val):
    """区分未提及、不适用、未识别三种状态。"""
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("", "空", "无", "null", "None", "/"):
        return ""
    if s in ("不适用", "N/A", "n/a"):
        return "不适用"
    if s in ("无法识别", "未识别"):
        return "未识别"
    return s


def normalize_info_type(val):
    """将旧的五分类统一收敛为两分类。"""
    normalized = normalize_text(val)
    if not normalized:
        return ""

    bidding_info_types = {
        "招标公告",
        "询比/竞价公告",
        "供应商征集公告",
        "招标信息",
    }
    result_info_types = {
        "中标/成交候选人公示",
        "中标/成交结果公告",
        "投标结果公告",
    }

    if normalized in bidding_info_types:
        return "招标信息"
    if normalized in result_info_types:
        return "投标结果公告"
    return normalized


def normalize_field_text(field_name: str, val):
    """按字段语义决定是否允许“不适用”落表。"""
    normalized = normalize_info_type(val) if field_name == "信息类型" else normalize_text(val)
    not_applicable_fields = {"开标时间", "中标单位", "中标金额（万元）"}
    if normalized == "不适用" and field_name not in not_applicable_fields:
        return ""
    return normalized


def display_cell_value(field_name: str, val):
    """统一表格中的占位符显示。"""
    normalized = normalize_field_text(field_name, val)
    if normalized in ("", "不适用", "未识别"):
        return "/"
    return normalized


def _format_wan(num):
    """将万元数字格式化为字符串"""
    if num == int(num):
        return str(int(num))
    return str(round(num, 4))


def extract_amount_from_text(text: str) -> str | None:
    """从金额文本中提取并换算为万元字符串。"""
    normalized = normalize_text(text)
    if not normalized or normalized in ("不适用", "未识别"):
        return None

    s = normalized.replace(",", "").replace("，", "")

    m = re.search(r'([\d.]+)\s*[（(]?\s*万(?:元)?[)）]?', s)
    if m:
        return _format_wan(float(m.group(1)))

    m = re.search(r'[¥￥]?\s*([\d.]+)\s*元', s)
    if m:
        return _format_wan(float(m.group(1)) / 10000)

    m = re.fullmatch(r'([\d.]+)', s)
    if m:
        num = float(m.group(1))
        if num >= 10000:
            return _format_wan(num / 10000)
        return _format_wan(num)

    return None


def normalize_amount(val, original_text=None):
    """
    统一金额为万元数字字符串。
    优先根据 original_text（原文）判断单位，再对 val 做校正。
    """
    normalized_val = normalize_text(val)
    normalized_original = normalize_text(original_text)

    if normalized_val == "不适用" or normalized_original == "不适用":
        return "不适用"
    if normalized_val == "未识别":
        return "未识别"

    parsed_original = extract_amount_from_text(normalized_original)
    if parsed_original is not None:
        return parsed_original

    if not normalized_val:
        return ""

    parsed_value = extract_amount_from_text(normalized_val)
    if parsed_value is not None:
        return parsed_value

    return normalized_val


def split_note_parts(note: str) -> list[str]:
    """按中文句号/分号拆分备注。"""
    normalized_note = normalize_text(note)
    if not normalized_note or normalized_note in ("不适用", "未识别"):
        return []
    parts = re.split(r"[；;。]\s*", normalized_note)
    return [part.strip(" ，,") for part in parts if part.strip(" ，,")]


def build_amount_note(label: str, amount_value: str, original_text) -> str:
    """为金额字段生成稳定、可核验的备注。"""
    normalized_original = normalize_text(original_text)

    if amount_value == "不适用":
        return ""
    if amount_value == "未识别":
        return f"{label}疑似存在但未能稳定识别"
    if not amount_value:
        if normalized_original:
            return f"{label}原文已提取但未能完成金额换算"
        return f"{label}未在文中提及"
    if normalized_original:
        return f"{label}按原文“{normalized_original}”换算为{amount_value}万元"
    return ""


def build_extract_note(data: dict, budget_value: str, bid_amount_value: str) -> str:
    """合并模型备注与程序生成的金额备注，避免互相矛盾。"""
    amount_keywords = ("预算金额", "中标金额", "成交金额", "金额原文", "换算为")
    note_parts = [
        part
        for part in split_note_parts(data.get("提取备注"))
        if not any(keyword in part for keyword in amount_keywords)
    ]

    def is_duplicate_part(part: str) -> bool:
        if budget_value == "" and ("预算字段未在文中提及" in part or "预算金额未在文中提及" in part):
            return True
        if bid_amount_value == "" and ("中标金额未在文中提及" in part or "成交金额未在文中提及" in part):
            return True
        info_type = normalize_info_type(data.get("信息类型"))
        if normalize_text(data.get("开标时间")) == "不适用" and "开标时间" in part and "不适用" in part:
            return True
        if normalize_text(data.get("中标单位")) == "不适用" and "中标单位" in part and "不适用" in part:
            return True
        if normalize_text(data.get("中标金额（万元）")) == "不适用" and "中标金额" in part and "不适用" in part:
            return True
        return False

    applicability_parts = []
    info_type = normalize_info_type(data.get("信息类型")) or "当前文件类型"
    for field_name, label in (
        ("开标时间", "开标时间"),
        ("中标单位", "中标单位"),
        ("中标金额（万元）", "中标金额"),
    ):
        state = normalize_text(data.get(field_name))
        if state == "不适用":
            applicability_parts.append(f"{label}不适用，因该文件为{info_type}")
        elif state == "未识别":
            applicability_parts.append(f"{label}未能稳定识别")

    generated_parts = [
        *applicability_parts,
        build_amount_note("预算金额", budget_value, data.get("预算原文")),
        build_amount_note("中标金额", bid_amount_value, data.get("中标金额原文")),
    ]

    merged = []
    for part in note_parts:
        if not is_duplicate_part(part) and part not in merged:
            merged.append(part)
    for part in generated_parts:
        if part and part not in merged:
            merged.append(part)
    return "；".join(merged)

def append_row(ws, filename: str, data: dict):
    """向工作表追加一行"""
    budget_value = normalize_amount(data.get("预算（万元）"), data.get("预算原文"))
    bid_amount_value = normalize_amount(data.get("中标金额（万元）"), data.get("中标金额原文"))

    row = [
        filename,
        display_cell_value("信息类型", data.get("信息类型")),
        display_cell_value("项目名称", data.get("项目名称")),
        display_cell_value("开标时间", data.get("开标时间")),
        display_cell_value("城市", data.get("城市")),
        display_cell_value("行业", data.get("行业")),
        display_cell_value("服务内容", data.get("服务内容")),
        "/" if budget_value in ("", "不适用", "未识别") else budget_value,
        display_cell_value("采购人", data.get("采购人")),
        display_cell_value("代理机构", data.get("代理机构")),
        display_cell_value("中标单位", data.get("中标单位")),
        "/" if bid_amount_value in ("", "不适用", "未识别") else bid_amount_value,
        build_extract_note(data, budget_value, bid_amount_value),
    ]
    ws.append(row)


# ──────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────

def process_pdfs(pdf_dir: str, output_excel: str, api_key: str, prompt: str):
    dashscope.api_key = api_key

    pdf_files = list(Path(pdf_dir).glob("*.pdf"))
    if not pdf_files:
        print(f"⚠️  在 {pdf_dir} 中未找到PDF文件")
        return

    print(f"📄 共找到 {len(pdf_files)} 个PDF文件")

    wb, ws = init_excel(output_excel)
    success, failed = 0, 0

    for i, pdf_path in enumerate(pdf_files, 1):
        filename = pdf_path.name
        print(f"\n[{i}/{len(pdf_files)}] 处理: {filename}")

        try:
            if is_image_based_pdf(str(pdf_path)):
                print("  → 图片型PDF，使用视觉模型")
                images_b64 = pdf_pages_to_base64(str(pdf_path))
                result = call_vision_model(prompt, images_b64)
            else:
                print("  → 文字型PDF，使用文本模型")
                text = extract_text_from_pdf(str(pdf_path))
                result = call_text_model(prompt, text)

            if result:
                append_row(ws, filename, result)
                wb.save(output_excel)
                print(f"  ✅ 成功 | 项目: {result.get('项目名称', '?')} | 类型: {result.get('信息类型', '?')}")
                success += 1
            else:
                append_row(ws, filename, {"项目名称": "解析失败"})
                wb.save(output_excel)
                print("  ❌ 模型返回解析失败")
                failed += 1

        except Exception as e:
            print(f"  ❌ 处理出错: {e}")
            append_row(ws, filename, {"项目名称": f"处理出错: {e}"})
            wb.save(output_excel)
            failed += 1

        # 避免API限速
        time.sleep(1)

    print(f"\n{'='*50}")
    print(f"✅ 成功: {success}  ❌ 失败: {failed}")
    print(f"📊 结果已保存至: {output_excel}")


def parse_args():
    parser = argparse.ArgumentParser(description="招标文件信息提取工具")
    parser.add_argument("--pdf_dir",      default="pdfs",        help="PDF文件目录")
    parser.add_argument("--excel_dir",    default="output",      help="输出Excel目录")
    parser.add_argument("--api_key_file", default="api_key.json",help="API密钥文件")
    parser.add_argument("--prompt_file",  default="prompt.txt",  help="提示词文件")
    return parser.parse_args()


if __name__ == "__main__":
    start = time.time()
    args = parse_args()

    # 路径转为绝对路径（兼容从任意目录调用）
    base_dir = Path(__file__).parent
    pdf_dir      = Path(args.pdf_dir) if Path(args.pdf_dir).is_absolute() else base_dir / args.pdf_dir
    excel_dir    = Path(args.excel_dir) if Path(args.excel_dir).is_absolute() else base_dir / args.excel_dir
    api_key_file = Path(args.api_key_file) if Path(args.api_key_file).is_absolute() else base_dir / args.api_key_file
    prompt_file  = Path(args.prompt_file) if Path(args.prompt_file).is_absolute() else base_dir / args.prompt_file

    # 创建输出目录
    excel_dir.mkdir(parents=True, exist_ok=True)
    output_excel = excel_dir / f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print("=" * 50)
    print("  BidMiner 招标文件信息提取工具")
    print("=" * 50)
    print(f"PDF目录:    {pdf_dir}")
    print(f"输出文件:   {output_excel}")

    api_key = load_api_key(str(api_key_file))
    prompt  = load_prompt(str(prompt_file))

    process_pdfs(str(pdf_dir), str(output_excel), api_key, prompt)

    print(f"⏱️  总耗时: {time.time() - start:.1f}s")
