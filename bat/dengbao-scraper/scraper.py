"""
项目进度数据爬虫
- 从项目管理系统爬取项目进度数据，导出为 Excel
- 支持多种项目类型：等保测评、密码评估、安全评估、风险评估、软件测试、安全服务、综合服务
- 支持自动登录（OCR 验证码）、手动指定 Cookie 两种鉴权方式

使用方法:
  python scraper.py                          # 默认爬取等保测评
  python scraper.py --type password          # 爬取密码评估
  python scraper.py --type all               # 爬取全部类型
  python scraper.py --cookie <PHPSESSID>     # 手动指定 cookie
  python scraper.py --output ./output        # 指定输出目录
  python scraper.py --limit 100              # 每页获取数量（默认 50）
"""

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============ 默认配置 ============
CONFIG = {
    "base_url": "",
    "pfx_path": "",
    "pfx_password": None,
    "username": "",
    "password": "",
    "cookie": "",
    "page_size": 50,
    "output_dir": "",
}

CONFIG_FILE = Path(__file__).parent / "config.json"

# ============ 项目类型定义 ============
PROJECT_TYPES = {
    "dengbao": {"name": "等保测评", "path": "/djcp/projectstatus/index"},
    "password": {"name": "密码评估", "path": "/smcp/projectstatus/index"},
    "security": {"name": "安全评估", "path": "/aqpg/projectstatus/index"},
    "risk": {"name": "风险评估", "path": "/fxpg/projectstatus/index"},
    "testing": {"name": "软件测试", "path": "/rjcs/projectstatus/index"},
    "service": {"name": "安全服务", "path": "/aqfw/projectstatus/index"},
    "comprehensive": {"name": "综合服务", "path": "/zhfw/projectstatus/index"},
}

# ============ Excel 列定义（字段映射） ============
COLUMNS = [
    ("系统编号", lambda r: r.get("system_id", "")),
    ("系统名称", lambda r: r.get("hand", {}).get("systemname", "")),
    ("客户名称", lambda r: r.get("hand", {}).get("customername", "")),
    ("系统级别", lambda r: r.get("hand", {}).get("systemlevel", "")),
    ("系统标签", lambda r: r.get("hand", {}).get("systemtag", "")),
    ("业务类型", lambda r: r.get("hand", {}).get("businesstype", "")),
    ("项目名称", lambda r: r.get("setup", {}).get("projectname", "")),
    ("项目编号", lambda r: r.get("setup", {}).get("project_id", "")),
    ("项目地点", lambda r: r.get("setup", {}).get("belongcity", "")),
    ("立项状态", lambda r: r.get("setup", {}).get("initstatus", "")),
    ("项目经理", lambda r: r.get("hand", {}).get("projectmanager", "")),
    ("项目部门", lambda r: r.get("hand", {}).get("pmdepartment", "")),
    ("销售负责人", lambda r: r.get("setup", {}).get("salewheel", "")),
    ("要求进场时间", lambda r: r.get("hand", {}).get("pstartdate", "")),
    ("要求完结时间", lambda r: r.get("hand", {}).get("pfinishdate", "")),
    ("实施开始日期", lambda r: r.get("startdate", "") or ""),
    ("实施结束日期", lambda r: r.get("finishdate", "") or ""),
    ("项目状态", lambda r: r.get("projectstatus_text", "")),
    ("是否完结", lambda r: r.get("isfinish_text", "")),
    ("方案打印", lambda r: r.get("isplanprint", "")),
    ("报告打印", lambda r: r.get("isreportprint", "")),
    ("备案状态", lambda r: r.get("hand", {}).get("isregister_text", "")),
    ("合同状态", lambda r: r.get("hand", {}).get("contractstatus_text", "")),
    ("备注", lambda r: r.get("remark", "") or ""),
]


# ============ 配置管理 ============

def load_config():
    """加载配置文件，与默认配置合并"""
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            saved = json.load(f)
            CONFIG.update(saved)
    return CONFIG


def save_config():
    """保存配置到文件"""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(CONFIG, f, ensure_ascii=False, indent=2)


# ============ PFX 证书处理 ============

def setup_client_cert(pfx_path, pfx_password=None):
    """从 PFX 提取 PEM 证书和私钥，返回临时文件路径"""
    from cryptography.hazmat.primitives.serialization import (
        pkcs12, Encoding, PrivateFormat, NoEncryption
    )

    with open(pfx_path, "rb") as f:
        pfx_data = f.read()

    private_key, certificate, ca_certs = pkcs12.load_key_and_certificates(
        pfx_data, pfx_password
    )

    cert_file = tempfile.NamedTemporaryFile(suffix=".pem", delete=False, mode="wb")
    key_file = tempfile.NamedTemporaryFile(suffix=".pem", delete=False, mode="wb")

    cert_file.write(certificate.public_bytes(Encoding.PEM))
    if ca_certs:
        for ca in ca_certs:
            cert_file.write(ca.public_bytes(Encoding.PEM))
    cert_file.close()

    key_file.write(
        private_key.private_bytes(
            Encoding.PEM, PrivateFormat.TraditionalOpenSSL, NoEncryption()
        )
    )
    key_file.close()

    return cert_file.name, key_file.name


# ============ 自动登录 ============

def _ensure_ddddocr():
    """确保 ddddocr 已安装"""
    try:
        import ddddocr
        return ddddocr
    except ImportError:
        print("[..] ddddocr 未安装，正在安装（首次较慢）...")
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "ddddocr", "-q"],
            check=True,
        )
        import ddddocr
        return ddddocr


def auto_login(session, base_url, username, password, max_retries=5):
    """
    自动登录：获取验证码 → OCR 识别 → 提交表单
    验证码识别可能失败，最多重试 max_retries 次
    返回: True 登录成功, False 失败
    """
    ddddocr = _ensure_ddddocr()
    ocr = ddddocr.DdddOcr(show_ad=False)

    login_url = f"{base_url}/index/login"
    captcha_base = base_url.rsplit("/", 1)[0]
    captcha_url = f"{captcha_base}/index.php?s=/captcha"

    for attempt in range(1, max_retries + 1):
        resp = session.get(login_url, timeout=15)
        token_match = re.search(r'name="__token__"\s+value="([^"]+)"', resp.text)
        token = token_match.group(1) if token_match else ""

        resp = session.get(captcha_url, timeout=10)
        captcha_text = ocr.classification(resp.content)
        print(f"  [..] 第 {attempt} 次尝试，识别验证码: {captcha_text}")

        login_data = {
            "__token__": token,
            "username": username,
            "password": password,
            "captcha": captcha_text,
            "keeplogin": "1",
        }
        resp = session.post(login_url, data=login_data, timeout=15)

        try:
            result = resp.json()
            if result.get("code") == 1:
                print(f"  [OK] 登录成功!")
                phpsessid = session.cookies.get("PHPSESSID")
                if phpsessid:
                    CONFIG["cookie"] = phpsessid
                    save_config()
                    print(f"  [OK] Cookie 已自动保存")
                return True
            else:
                msg = result.get("msg", "未知错误")
                print(f"  [!] 登录失败: {msg}")
        except ValueError:
            if "登录成功" in resp.text or resp.status_code == 302:
                print(f"  [OK] 登录成功!")
                phpsessid = session.cookies.get("PHPSESSID")
                if phpsessid:
                    CONFIG["cookie"] = phpsessid
                    save_config()
                return True
            err_match = re.search(r'<div[^>]*class="content"[^>]*>\s*<p>(.*?)</p>', resp.text)
            if err_match:
                print(f"  [!] 登录失败: {err_match.group(1)}")
            else:
                print(f"  [!] 登录失败（HTML 响应）")

        time.sleep(0.5)

    # 所有重试都失败后，检查是否实际已处于登录状态
    # （已登录时访问登录页会返回非预期的 HTML，导致误判为失败）
    if check_session(session, base_url):
        print("  [OK] 检测到已登录状态")
        phpsessid = session.cookies.get("PHPSESSID")
        if phpsessid:
            CONFIG["cookie"] = phpsessid
            save_config()
        return True

    print(f"[X] 登录失败，已尝试 {max_retries} 次")
    return False


def check_session(session, base_url, api_path="/djcp/projectstatus/index"):
    """检查当前 session 是否有效"""
    url = f"{base_url}{api_path}"
    params = {"offset": "0", "limit": "1", "filter": "{}", "op": "{}"}
    headers = {"X-Requested-With": "XMLHttpRequest"}
    try:
        resp = session.get(url, headers=headers, params=params, timeout=10)
        data = resp.json()
        if "total" in data:
            return True
        if data.get("msg") == "请登录后操作":
            return False
    except Exception:
        pass
    return False


# ============ Session 创建 ============

def create_session(pfx_path, pfx_password=None, cookie=None):
    """创建带证书的 requests Session，可选附加 Cookie"""
    cert_path, key_path = setup_client_cert(pfx_path, pfx_password)

    session = requests.Session()
    session.verify = False
    session.cert = (cert_path, key_path)
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/146.0.0.0 Safari/537.36"
        ),
    })

    if cookie:
        session.cookies.set("PHPSESSID", cookie)

    return session, cert_path, key_path


# ============ 数据获取 ============

def fetch_page(session, base_url, api_path, offset, limit):
    """获取一页数据"""
    url = f"{base_url}{api_path}"
    params = {
        "addtabs": "1",
        "sort": "id",
        "order": "desc",
        "offset": str(offset),
        "limit": str(limit),
        "filter": "{}",
        "op": "{}",
    }
    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json",
        "X-Requested-With": "XMLHttpRequest",
    }
    resp = session.get(url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if "code" in data and data.get("msg") == "请登录后操作":
        raise RuntimeError("Session 已过期")

    return data


def fetch_all(session, base_url, api_path, page_size=50):
    """分页获取全部数据"""
    first_page = fetch_page(session, base_url, api_path, 0, page_size)
    total = first_page.get("total", 0)
    rows = first_page.get("rows", [])

    print(f"[OK] 总记录数: {total}")
    print(f"[OK] 第 1 页: 获取 {len(rows)} 条")

    offset = page_size
    page_num = 2
    while offset < total:
        time.sleep(0.3)
        data = fetch_page(session, base_url, api_path, offset, page_size)
        page_rows = data.get("rows", [])
        rows.extend(page_rows)
        print(f"[OK] 第 {page_num} 页: 获取 {len(page_rows)} 条 (累计 {len(rows)}/{total})")
        offset += page_size
        page_num += 1

    print(f"[OK] 完成，共获取 {len(rows)} 条记录")
    return rows


# ============ Excel 导出 ============

def export_excel(rows, output_dir, type_name="等保测评"):
    """导出数据到 Excel"""
    try:
        import openpyxl
    except ImportError:
        print("[..] openpyxl 未安装，正在安装...")
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            check=True,
        )
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type_name

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_, extractor) in enumerate(COLUMNS, 1):
            value = extractor(row_data)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(rows) + 2, 52)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_path / f"{type_name}_{timestamp}.xlsx"
    wb.save(str(filename))

    print(f"[OK] 已导出: {filename}")
    print(f"[OK] 共 {len(rows)} 条记录, {len(COLUMNS)} 列")
    return str(filename)


# ============ 单类型爬取流程 ============

def run_scrape(session, base_url, api_path, type_name, output_dir, page_size):
    """执行单个类型的爬取流程（已登录状态）"""
    print(f"\n[..] 开始获取数据: {type_name}")
    rows = fetch_all(session, base_url, api_path, page_size)

    print("[..] 导出 Excel...")
    filepath = export_excel(rows, output_dir, type_name)
    return filepath, len(rows)


# ============ 主流程 ============

def main():
    parser = argparse.ArgumentParser(description="项目进度数据爬虫")
    parser.add_argument("--type", default="dengbao",
                        choices=list(PROJECT_TYPES.keys()) + ["all"],
                        help="项目类型（默认 dengbao），使用 all 爬取全部类型")
    parser.add_argument("--cookie", help="手动指定 PHPSESSID Cookie（跳过自动登录）")
    parser.add_argument("--username", help="登录账号")
    parser.add_argument("--password", help="登录密码")
    parser.add_argument("--pfx", help="PFX 证书文件路径")
    parser.add_argument("--url", help="系统 base URL")
    parser.add_argument("--output", help="输出目录")
    parser.add_argument("--limit", type=int, help="每页数量（默认 50）")
    args = parser.parse_args()

    load_config()

    # 命令行参数覆盖配置
    if args.username:
        CONFIG["username"] = args.username
    if args.password:
        CONFIG["password"] = args.password
    if args.pfx:
        CONFIG["pfx_path"] = args.pfx
    if args.url:
        CONFIG["base_url"] = args.url

    output_dir = args.output or CONFIG["output_dir"] or str(Path(__file__).parent / "output")
    page_size = args.limit or CONFIG["page_size"]

    # 确定要爬取的类型
    if args.type == "all":
        scrape_types = list(PROJECT_TYPES.keys())
        type_label = "全部类型"
    else:
        scrape_types = [args.type]
        type_label = PROJECT_TYPES[args.type]["name"]

    print("=" * 55)
    print(f"  项目进度数据爬虫 — {type_label}")
    print("=" * 55)
    print()

    # 创建 session
    print("[..] 初始化连接...")
    cookie = args.cookie or CONFIG.get("cookie", "")
    session, cert_path, key_path = create_session(
        CONFIG["pfx_path"], CONFIG["pfx_password"], cookie or None
    )

    try:
        # 鉴权流程: 先试已有 cookie → 失效则自动登录
        # 用等保路径检查 session（PHPSESSID 是全局的，与具体模块无关）
        if cookie:
            print("[..] 检查已有 Session...")
            if check_session(session, CONFIG["base_url"]):
                print("[OK] Session 有效")
            else:
                print("[!] Session 已过期，尝试自动登录...")
                cookie = None

        if not cookie:
            print("[..] 自动登录中...")
            success = auto_login(
                session, CONFIG["base_url"],
                CONFIG["username"], CONFIG["password"],
            )
            if not success:
                print("[X] 自动登录失败，请手动指定 Cookie:")
                print("    python scraper.py --cookie <PHPSESSID>")
                sys.exit(1)

        # 逐个类型爬取
        total_records = 0
        results = []

        for type_key in scrape_types:
            type_info = PROJECT_TYPES[type_key]
            type_name = type_info["name"]
            api_path = type_info["path"]

            try:
                filepath, count = run_scrape(
                    session, CONFIG["base_url"], api_path,
                    type_name, output_dir, page_size,
                )
                total_records += count
                results.append((type_name, count, filepath))
            except Exception as e:
                print(f"\n[X] {type_name} 爬取失败: {e}")
                results.append((type_name, 0, None))

        # 汇总
        print()
        print("=" * 55)
        if len(results) > 1:
            print("  爬取汇总:")
            for name, count, fpath in results:
                status = f"{count} 条" if fpath else "失败"
                print(f"    {name}: {status}")
            print(f"  总计: {total_records} 条记录")
        else:
            name, count, fpath = results[0]
            print(f"完成! {name}: {count} 条记录")
            if fpath:
                print(f"文件: {fpath}")
        print("=" * 55)

    except RuntimeError as e:
        print(f"\n[X] {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n[X] 发生错误: {e}")
        raise
    finally:
        for f in (cert_path, key_path):
            try:
                os.unlink(f)
            except OSError:
                pass


if __name__ == "__main__":
    main()
