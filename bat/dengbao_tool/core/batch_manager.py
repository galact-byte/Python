"""批量项目/系统扫描与状态持久化。"""

from __future__ import annotations

import glob
import hashlib
import json
import os
from datetime import datetime
from typing import Dict, List, Tuple

STATE_FILENAME = "batch_update_state.json"
SCAN_EXCLUDE_DIRS = {
    "__pycache__",
    ".git",
    ".idea",
    ".vscode",
    ".omc",
    "doc_templates",
    "output",
}


def scan_batch_root(root_dir: str) -> List[dict]:
    """扫描总目录，返回项目列表。"""
    if not root_dir or not os.path.isdir(root_dir):
        return []

    projects: List[dict] = []
    child_dirs = [
        os.path.join(root_dir, name)
        for name in sorted(os.listdir(root_dir))
        if os.path.isdir(os.path.join(root_dir, name))
        and name not in SCAN_EXCLUDE_DIRS
        and not name.startswith(".")
    ]

    for child_dir in child_dirs:
        project = scan_project(child_dir)
        if project["systems"]:
            projects.append(project)

    if not projects:
        fallback = scan_project(root_dir)
        if fallback["systems"]:
            projects.append(fallback)

    return projects


def scan_project(project_dir: str) -> dict:
    """扫描单个项目目录。"""
    state = load_project_state(project_dir)
    systems = _discover_systems(project_dir, state)

    pending_count = sum(1 for item in systems if item.get("needs_update"))
    updated_count = sum(1 for item in systems if not item.get("needs_update"))
    project_name = state.get("project_name") or _guess_project_name(project_dir, systems)

    return {
        "project_dir": project_dir,
        "project_name": project_name,
        "output_dir": state.get("output_dir") or project_dir,
        "state_path": os.path.join(project_dir, STATE_FILENAME),
        "pending_count": pending_count,
        "updated_count": updated_count,
        "system_count": len(systems),
        "systems": systems,
    }


def load_project_state(project_dir: str) -> dict:
    """读取项目状态文件。"""
    state_path = os.path.join(project_dir, STATE_FILENAME)
    if not os.path.exists(state_path):
        return {
            "version": 1,
            "project_name": "",
            "output_dir": project_dir,
            "systems": {},
        }

    try:
        with open(state_path, "r", encoding="utf-8") as f:
            state = json.load(f)
    except Exception:
        state = {}

    if not isinstance(state, dict):
        state = {}

    state.setdefault("version", 1)
    state.setdefault("project_name", "")
    state.setdefault("output_dir", project_dir)
    state.setdefault("systems", {})
    return state


def save_project_state(project_dir: str, state: dict) -> str:
    """保存项目状态文件。"""
    state = dict(state or {})
    state.setdefault("version", 1)
    state.setdefault("project_name", "")
    state.setdefault("output_dir", project_dir)
    state.setdefault("systems", {})
    state["updated_at"] = _now_str()

    state_path = os.path.join(project_dir, STATE_FILENAME)
    with open(state_path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    return state_path


def get_system_entry(project_dir: str, system_id: str) -> Tuple[dict, dict]:
    """获取项目状态及系统条目。"""
    state = load_project_state(project_dir)
    entry = state.get("systems", {}).get(system_id, {})
    return state, entry


def upsert_system_entry(
    state: dict,
    project_dir: str,
    system_meta: dict,
    *,
    project_name: str = "",
    output_dir: str = "",
    form_data: dict | None = None,
    report_data: dict | None = None,
    ui_state: dict | None = None,
) -> dict:
    """更新系统状态条目。"""
    state.setdefault("systems", {})
    system_id = system_meta["system_id"]
    entry = dict(state["systems"].get(system_id, {}))

    signature = build_source_signature(system_meta)
    entry.update(
        {
            "system_id": system_id,
            "system_name": system_meta.get("system_name") or entry.get("system_name") or "",
            "source_dir": system_meta.get("source_dir") or entry.get("source_dir") or project_dir,
            "old_beian": system_meta.get("old_beian") or entry.get("old_beian") or "",
            "old_report": system_meta.get("old_report") or entry.get("old_report") or "",
            "survey": system_meta.get("survey") or entry.get("survey") or "",
            "output_dir": output_dir or system_meta.get("output_dir") or entry.get("output_dir") or system_meta.get("source_dir") or project_dir,
            "source_signature": signature,
            "source_signature_digest": signature_digest(signature),
            "updated_at": _now_str(),
        }
    )

    if form_data is not None:
        entry["form_data"] = form_data
    if report_data is not None:
        entry["report_data"] = report_data
    if ui_state is not None:
        entry["ui_state"] = ui_state
    if form_data is not None or report_data is not None:
        entry["draft_hash"] = calc_data_hash(
            entry.get("form_data", {}),
            entry.get("report_data", {}),
        )

    if project_name:
        state["project_name"] = project_name
    state["output_dir"] = output_dir or state.get("output_dir") or project_dir
    state["systems"][system_id] = entry
    return entry


def mark_generated(state: dict, system_id: str, generated_files: List[str]) -> dict:
    """标记系统已生成。"""
    entry = state.setdefault("systems", {}).setdefault(system_id, {})
    entry["generated_files"] = generated_files
    entry["generated_at"] = _now_str()
    entry["generated_hash"] = entry.get("draft_hash") or calc_data_hash(
        entry.get("form_data", {}),
        entry.get("report_data", {}),
    )
    return entry


def build_source_signature(system_meta: dict) -> dict:
    """构建源文件签名。"""
    signature = {}
    for key in ("old_beian", "old_report", "survey"):
        path = system_meta.get(key) or ""
        if path and os.path.exists(path):
            stat = os.stat(path)
            signature[key] = {
                "path": path,
                "mtime": int(stat.st_mtime),
                "size": stat.st_size,
            }
    return signature


def signature_digest(signature: dict) -> str:
    raw = json.dumps(signature or {}, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def calc_data_hash(form_data: dict, report_data: dict) -> str:
    raw = json.dumps(
        {"form_data": form_data or {}, "report_data": report_data or {}},
        ensure_ascii=False,
        sort_keys=True,
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def import_manifest(manifest_path: str) -> dict:
    """读取 Excel 清单，按项目分组。"""
    from openpyxl import load_workbook

    wb = load_workbook(manifest_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"projects": [], "sheet_name": ws.title}

    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def value(row, header_name):
        idx = header_map.get(header_name)
        if idx is None or idx >= len(row):
            return ""
        item = row[idx]
        return "" if item is None else str(item).strip()

    grouped: Dict[str, dict] = {}
    for row in rows[1:]:
        project_name = value(row, "项目名称")
        if not project_name:
            continue
        project = grouped.setdefault(
            project_name,
            {
                "project_name": project_name,
                "project_code": value(row, "项目编号"),
                "customer_name": value(row, "客户名称"),
                "systems": [],
            },
        )
        project["systems"].append(
            {
                "system_name": value(row, "系统名称"),
                "system_code": value(row, "系统编号"),
                "level": value(row, "系统级别"),
                "tag": value(row, "系统标签"),
                "beian_status": value(row, "备案状态"),
            }
        )

    return {
        "sheet_name": ws.title,
        "projects": list(grouped.values()),
    }


def _discover_systems(project_dir: str, state: dict) -> List[dict]:
    systems_by_id: Dict[str, dict] = {}
    for source_dir in _iter_candidate_dirs(project_dir):
        system_meta = _build_system_meta(project_dir, source_dir)
        if not system_meta:
            continue
        systems_by_id[system_meta["system_id"]] = _merge_system_meta(
            system_meta,
            state.get("systems", {}).get(system_meta["system_id"], {}),
        )

    for system_id, entry in state.get("systems", {}).items():
        if system_id in systems_by_id:
            continue
        systems_by_id[system_id] = _merge_system_meta(
            {
                "system_id": system_id,
                "system_name": entry.get("system_name") or system_id,
                "source_dir": entry.get("source_dir") or project_dir,
                "old_beian": entry.get("old_beian", ""),
                "old_report": entry.get("old_report", ""),
                "survey": entry.get("survey", ""),
                "output_dir": entry.get("output_dir") or entry.get("source_dir") or project_dir,
            },
            entry,
        )

    return sorted(systems_by_id.values(), key=lambda item: item.get("system_name", ""))


def _merge_system_meta(system_meta: dict, entry: dict) -> dict:
    signature = build_source_signature(system_meta)
    digest = signature_digest(signature)
    generated_files = entry.get("generated_files", [])
    generated_files_exist = bool(generated_files) and all(os.path.exists(path) for path in generated_files)
    draft_hash = entry.get("draft_hash") or calc_data_hash(
        entry.get("form_data", {}),
        entry.get("report_data", {}),
    )
    generated_hash = entry.get("generated_hash", "")
    source_changed = bool(entry) and entry.get("source_signature_digest") not in ("", digest)
    has_source = bool(system_meta.get("old_beian") or system_meta.get("old_report"))
    has_saved_data = bool(entry.get("form_data") or entry.get("report_data"))
    needs_update = (
        not generated_files_exist
        or not generated_hash
        or generated_hash != draft_hash
        or source_changed
    )
    status = "pending" if needs_update else "updated"
    if not has_source and not has_saved_data:
        status = "missing_source"
        needs_update = False

    merged = {
        "system_id": system_meta["system_id"],
        "system_name": entry.get("system_name") or system_meta.get("system_name") or system_meta["system_id"],
        "source_dir": system_meta.get("source_dir", ""),
        "old_beian": system_meta.get("old_beian", ""),
        "old_report": system_meta.get("old_report", ""),
        "survey": system_meta.get("survey", ""),
        "output_dir": entry.get("output_dir") or system_meta.get("output_dir") or system_meta.get("source_dir", ""),
        "generated_at": entry.get("generated_at", ""),
        "generated_files": generated_files,
        "generated_files_exist": generated_files_exist,
        "status": status,
        "needs_update": needs_update,
        "has_saved_data": has_saved_data,
        "source_changed": source_changed,
        "source_signature_digest": digest,
    }
    return merged


def _build_system_meta(project_dir: str, source_dir: str) -> dict | None:
    old_beian = _find_file(source_dir, "*备案表*.docx") or _find_file(source_dir, "*备案表*.doc")
    old_report = _find_file(source_dir, "*定级报告*.docx") or _find_file(source_dir, "*定级报告*.doc")
    survey = (
        _find_file(source_dir, "*调查表*.docx")
        or _find_file(source_dir, "*基本情况*.docx")
        or _find_file(os.path.dirname(source_dir), "*调查表*.docx")
        or _find_file(os.path.dirname(source_dir), "*基本情况*.docx")
    )
    if not old_beian and not old_report:
        return None

    system_name = _guess_system_name(source_dir, old_beian, old_report)
    system_id = _build_system_id(project_dir, source_dir)
    return {
        "system_id": system_id,
        "system_name": system_name,
        "source_dir": source_dir,
        "old_beian": old_beian or "",
        "old_report": old_report or "",
        "survey": survey or "",
        "output_dir": source_dir,
    }


def _iter_candidate_dirs(project_dir: str, max_depth: int = 2):
    for root, dirs, _files in os.walk(project_dir):
        rel_path = os.path.relpath(root, project_dir)
        depth = 0 if rel_path == "." else rel_path.count(os.sep) + 1
        dirs[:] = [
            name
            for name in dirs
            if name not in SCAN_EXCLUDE_DIRS and not name.startswith(".")
        ]
        if depth > max_depth:
            dirs[:] = []
            continue
        yield root


def _guess_project_name(project_dir: str, systems: List[dict]) -> str:
    if systems:
        primary = systems[0].get("system_name", "")
        if primary:
            return os.path.basename(project_dir) if len(systems) > 1 else primary
    return os.path.basename(project_dir)


def _guess_system_name(source_dir: str, old_beian: str, old_report: str) -> str:
    for path in (old_beian, old_report):
        if not path:
            continue
        stem = os.path.splitext(os.path.basename(path))[0]
        for prefix in ("备案表_", "定级报告_", "预览_备案表_", "预览_定级报告_", "01-新备案表"):
            if stem.startswith(prefix):
                stem = stem[len(prefix):].strip()
        if stem:
            return stem
    return os.path.basename(source_dir)


def _build_system_id(project_dir: str, source_dir: str) -> str:
    rel_path = os.path.relpath(source_dir, project_dir).replace("\\", "/")
    rel_path = "root" if rel_path == "." else rel_path
    safe = "".join(ch if ch.isalnum() or ch in ("-", "_", "/") else "_" for ch in rel_path)
    suffix = hashlib.sha1(rel_path.encode("utf-8")).hexdigest()[:8]
    return f"{safe}__{suffix}"


def _find_file(directory: str, pattern: str) -> str | None:
    if not directory or not os.path.isdir(directory):
        return None
    matches = glob.glob(os.path.join(directory, pattern))
    return min(matches, key=len) if matches else None


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
